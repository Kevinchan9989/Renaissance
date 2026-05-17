"""Audit unique fill colors in the workbook (find any TBC yellows we missed)."""
import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

PATH = r"C:\Users\ECQ1025\Downloads\MAS Securities\MAS DM\Renaissance\gsib-migration\workspace\MAS_OMEGA_DM_Conversion_Mapping_Specification _v0.01.xlsx"
wb = openpyxl.load_workbook(PATH, data_only=True)

seen = {}
for sn in wb.sheetnames:
    ws = wb[sn]
    for row in ws.iter_rows():
        for cell in row:
            f = cell.fill
            if f and f.fgColor:
                rgb = getattr(f.fgColor, "rgb", None)
                if rgb and isinstance(rgb, str) and rgb not in {"00000000", "FFFFFFFF"}:
                    key = (sn, rgb)
                    if key not in seen:
                        seen[key] = (cell.coordinate, str(cell.value)[:60] if cell.value else "")

for (sn, rgb), (coord, val) in sorted(seen.items()):
    print(f"{sn:30s} {rgb}  @ {coord}  {val!r}")
