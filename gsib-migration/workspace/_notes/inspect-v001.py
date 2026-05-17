"""
Inspect MAS_OMEGA_DM_Conversion_Mapping_Specification _v0.01.xlsx

Dumps every sheet's content + style metadata (yellow fills for TBC) so we can
build a faithful style guide / compare against tools/conversion-spec-exporter/data/sample.cjs.
"""

import openpyxl
from openpyxl.styles import PatternFill
import json
import sys
import io

# Force UTF-8 stdout on Windows
sys.stdout.reconfigure(encoding='utf-8')

PATH = r"C:\Users\ECQ1025\Downloads\MAS Securities\MAS DM\Renaissance\gsib-migration\workspace\MAS_OMEGA_DM_Conversion_Mapping_Specification _v0.01.xlsx"

wb = openpyxl.load_workbook(PATH, data_only=True)

print("=" * 100)
print("WORKBOOK SHEETS")
print("=" * 100)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"  - {sn!r:50s} (rows={ws.max_row}, cols={ws.max_column})")

print()


def cell_repr(c):
    if c is None:
        return ""
    v = c if not hasattr(c, "value") else c.value
    if v is None:
        return ""
    s = str(v)
    return s


def is_yellow(cell):
    """Detect yellow fill (TBC marker)."""
    if cell.fill is None:
        return False
    fg = cell.fill.fgColor
    if fg is None:
        return False
    rgb = getattr(fg, "rgb", None)
    if rgb is None:
        return False
    if not isinstance(rgb, str):
        return False
    rgb = rgb.upper()
    # Common yellows: FFFF00, FFFFFF00, FFFFE699, etc.
    if rgb in {"FFFFFF00", "00FFFF00", "FFFFFFE0", "FFFFFFCC", "FFFFFF99", "FFFFE699"}:
        return True
    # Heuristic: high R and G, low B
    if len(rgb) == 8:
        try:
            r = int(rgb[2:4], 16)
            g = int(rgb[4:6], 16)
            b = int(rgb[6:8], 16)
            if r >= 200 and g >= 200 and b < 200:
                return True
        except ValueError:
            return False
    return False


def dump_sheet(sn, max_rows=None):
    ws = wb[sn]
    print("=" * 100)
    print(f"SHEET: {sn!r}    (rows={ws.max_row}, cols={ws.max_column})")
    print("=" * 100)
    for i, row in enumerate(ws.iter_rows(), 1):
        if max_rows and i > max_rows:
            print(f"... (truncated at row {max_rows})")
            break
        # Skip purely empty
        vals = [cell_repr(c) for c in row]
        if not any(v.strip() for v in vals):
            continue
        # Detect yellow fills
        yellows = [j for j, c in enumerate(row, 1) if is_yellow(c)]
        flag = f"  [YELLOW cols={yellows}]" if yellows else ""
        # Pretty print
        joined = " | ".join(v.replace("\n", "\\n") for v in vals)
        print(f"r{i:03d}: {joined}{flag}")
    print()


# Dump every sheet
for sn in wb.sheetnames:
    dump_sheet(sn)
